from io import BytesIO
import unittest

from openpyxl import Workbook

from core.catalog import build_catalog_from_files


def workbook_bytes(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class CatalogTests(unittest.TestCase):
    def test_parses_trd_hierarchy_and_retention(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "TRD JURIDICA"
        ws["B1"] = "TABLA DE RETENCIÓN DOCUMENTAL - TRD"
        ws["B4"] = "Dependencia Productora:"
        ws["C4"] = "JURÍDICA"
        ws["I4"] = "Código:"
        ws["K4"] = 1004
        ws["A5"] = "CÓDIGO"
        ws["B5"] = "SERIE - SUBSERIE Y TIPOS DOCUMENTAL"
        ws["C6"] = "Papel"
        ws["D6"] = "Electrónico (extensión)"
        ws["E6"] = "Archivo de Gestión"
        ws["F6"] = "Archivo Central"
        ws["G6"] = "C"
        ws["H6"] = "S"
        ws["I6"] = "E"
        ws["J6"] = "D"
        ws["K5"] = "PROCEDIMIENTO"
        ws["A7"] = "1004.13.31"
        ws["B7"] = "CONTRATOS"
        ws["E7"] = 3
        ws["F7"] = 17
        ws["I7"] = "X"
        ws["K7"] = "Eliminar después de la retención aprobada."
        ws["B8"] = "Contratos de Arrendamiento"
        ws["B9"] = "Contrato"
        ws["C9"] = "X"
        ws["B10"] = "Póliza"

        result = build_catalog_from_files([("1004 TRD JURIDICA.xlsx", workbook_bytes(wb))])

        self.assertEqual(result.errors, [])
        self.assertEqual(len(result.entries), 1)
        entry = result.entries[0]
        self.assertEqual(entry["code"], "1004.13.31")
        self.assertEqual(entry["series_name"], "CONTRATOS")
        self.assertEqual(entry["subseries_name"], "Contratos de Arrendamiento")
        self.assertEqual(entry["document_types"], ["Contrato", "Póliza"])
        self.assertEqual(entry["retention_management"], "3")
        self.assertEqual(entry["final_disposition"], ["Eliminación"])

    def test_ccd_enriches_trd_names(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "C.C.D."
        ws["B1"] = "CUADRO DE CLASIFICACIÓN DOCUMENTAL - CCD"
        headers = ["CÓDIGO", "SECCIÓN", "CÓDIGO", "SUBSECCIÓN", "CÓDIGO", "SERIES", "CÓDIGO", "SUBSERIES"]
        for col, value in enumerate(headers, 1):
            ws.cell(5, col, value)
        values = [1000, "GERENCIA GENERAL", 1004, "JURÍDICA", 13, "CONTRATOS", 31, "Contratos de Arrendamiento"]
        for col, value in enumerate(values, 1):
            ws.cell(6, col, value)

        result = build_catalog_from_files([("CCD.xlsx", workbook_bytes(wb))])

        self.assertEqual(len(result.entries), 1)
        self.assertEqual(result.entries[0]["code"], "1004.13.31")
        self.assertEqual(result.entries[0]["dependency_name"], "JURÍDICA")


if __name__ == "__main__":
    unittest.main()
