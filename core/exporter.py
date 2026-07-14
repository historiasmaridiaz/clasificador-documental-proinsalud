from __future__ import annotations

import csv
import io
import json
import re
import unicodedata
import zipfile
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path, PurePosixPath
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .hierarchy import canonicalize_custom_entry


EXPORT_COLUMNS = [
    "anio",
    "archivo",
    "codigo_sugerido",
    "puntaje_similitud",
    "confianza",
    "codigo_final",
    "estado",
    "dependencia",
    "serie",
    "subserie",
    "tipo_documental",
    "retencion_gestion",
    "retencion_central",
    "disposicion_final",
    "fuente_clasificacion",
    "origen_anio",
    "observaciones",
]


OTHER_SERIES_LABEL = "OTROS / OTRAS SERIES — sin código TRD"
OTHER_SERIES_FOLDER = "OTROS - OTRAS SERIES SIN CÓDIGO TRD"
OTHER_DEPENDENCY_FOLDER = "OTROS - DEPENDENCIA POR DEFINIR"


def _generic_other_entry(record: dict[str, Any]) -> dict[str, Any]:
    """Crea una clasificación temporal segura para registros fuera de la TRD.

    Se conserva, cuando existe, la dependencia propuesta por la IA para que el
    documento quede dentro del área correcta. No se inventan códigos, series ni
    subseries oficiales.
    """

    candidate: dict[str, Any] = {}
    candidates = record.get("candidates") or []
    if isinstance(candidates, list) and candidates and isinstance(candidates[0], dict):
        candidate = dict(candidates[0])
    dependency_code = str(candidate.get("dependency_code") or record.get("dependency_code") or "").strip()
    dependency_name = str(candidate.get("dependency_name") or record.get("dependency_name") or "").strip()
    return {
        "is_other": True,
        "is_banter": False,
        "source": "OTROS / sin coincidencia en la TRD activa",
        "code": "",
        "dependency_code": dependency_code,
        "dependency_name": dependency_name,
        "series_code": "",
        "series_name": OTHER_SERIES_LABEL,
        "subseries_code": "",
        "subseries_name": "",
        "document_types": [],
        "retention_management": "",
        "retention_central": "",
        "final_disposition": [],
        "procedure": "Clasificación temporal pendiente de revisión archivística.",
    }


def _catalog_map(entries: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {str(entry.get("code", "")): entry for entry in entries if entry.get("code")}


def _effective_entry(record: dict[str, Any], catalog: dict[str, dict[str, Any]]) -> dict[str, Any]:
    custom = record.get("custom_classification") or {}
    if isinstance(custom, dict) and custom:
        return dict(custom)
    code = str(record.get("final_code") or record.get("suggested_code") or "")
    if code and code in catalog:
        return dict(catalog[code])

    # Un código ausente o que ya no pertenece al catálogo activo no debe caer
    # en una carpeta paralela SIN_CLASIFICACION. Se conserva en la serie
    # temporal OTROS, tal como se presenta en la interfaz de revisión.
    return _generic_other_entry(record)


def manifest_rows(records: list[dict[str, Any]], entries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    catalog = _catalog_map(entries)
    rows = []
    for record in records:
        entry = canonicalize_custom_entry(
            _effective_entry(record, catalog),
            entries,
            filename=str(record.get("filename") or ""),
            extracted_text=str(record.get("extracted_text") or ""),
        )
        code = "" if entry.get("is_other") else str(entry.get("code") or record.get("final_code") or record.get("suggested_code") or "")
        document_types = entry.get("document_types") or []
        if isinstance(document_types, str):
            document_types = [document_types]
        rows.append(
            {
                "anio": record.get("document_year", ""),
                "archivo": record.get("filename", ""),
                "codigo_sugerido": record.get("suggested_code", ""),
                "puntaje_similitud": round(float(record.get("suggested_score", 0)) * 100, 1),
                "confianza": record.get("confidence", ""),
                "codigo_final": code,
                "estado": record.get("status", ""),
                "dependencia": entry.get("dependency_name", ""),
                "serie": entry.get("series_name", ""),
                "subserie": entry.get("subseries_name", ""),
                "tipo_documental": " | ".join(str(value) for value in document_types),
                "retencion_gestion": entry.get("retention_management", ""),
                "retencion_central": entry.get("retention_central", ""),
                "disposicion_final": " | ".join(entry.get("final_disposition") or []),
                "fuente_clasificacion": entry.get("source") or ("TRD PROINSALUD" if code else "Sin clasificación"),
                "origen_anio": record.get("year_source", ""),
                "observaciones": record.get("reviewer_notes", ""),
            }
        )
    return rows


def to_csv_bytes(rows: list[dict[str, Any]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=EXPORT_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return buffer.getvalue().encode("utf-8-sig")


def to_json_bytes(rows: list[dict[str, Any]]) -> bytes:
    return json.dumps(rows, ensure_ascii=False, indent=2).encode("utf-8")


def to_xlsx_bytes(rows: list[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Clasificación"
    ws.append(EXPORT_COLUMNS)
    for row in rows:
        ws.append([row.get(column, "") for column in EXPORT_COLUMNS])

    header_fill = PatternFill("solid", fgColor="123F73")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = [12, 35, 18, 18, 13, 18, 14, 28, 28, 34, 34, 18, 18, 28, 30, 28, 35]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1, index).column_letter].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _folder_component(value: Any, fallback: str, max_length: int = 120) -> str:
    text = unicodedata.normalize("NFC", str(value or fallback))
    text = re.sub(r'[<>:"/\\|?*\x00-\x1F]+', "_", text).strip(" ._")
    return text[:max_length].rstrip(" ._") or fallback


def _file_component(value: Any, fallback: str, max_length: int = 100) -> str:
    cleaned = _folder_component(value, fallback, max_length=500)
    suffix = Path(cleaned).suffix
    stem = Path(cleaned).stem
    available = max(12, max_length - len(suffix))
    return f"{stem[:available].rstrip(' ._')}{suffix}" or fallback


def _fit_zip_target(
    parts: list[str],
    filename: str,
    max_length: int = 240,
    preserve_parts: bool = False,
) -> tuple[list[str], str]:
    """Ajusta la ruta sin crear una segunda carpeta por truncamiento.

    Cuando la ruta proviene de la plantilla masiva, los nombres de carpeta se
    conservan exactamente. En ese caso se recorta únicamente el nombre del
    archivo. Esto evita que una carpeta vacía de la plantilla y una carpeta
    truncada para el documento aparezcan como dos carpetas casi idénticas.
    """
    fitted_parts = list(parts)
    fitted_file = filename

    def current_length() -> int:
        return len("/".join(fitted_parts + ([fitted_file] if fitted_file else [])))

    if current_length() <= max_length:
        return fitted_parts, fitted_file

    suffix = Path(fitted_file).suffix
    stem = Path(fitted_file).stem
    directory_length = len("/".join(fitted_parts)) + (1 if fitted_parts and fitted_file else 0)
    allowed_file_length = max(12, max_length - directory_length)
    if fitted_file and len(fitted_file) > allowed_file_length:
        available_stem = max(4, allowed_file_length - len(suffix))
        fitted_file = f"{stem[:available_stem].rstrip(' ._')}{suffix}"

    if preserve_parts:
        # Si la propia plantilla supera el límite conservador, se mantiene la
        # jerarquía oficial y no se genera una carpeta alternativa truncada.
        return fitted_parts, fitted_file

    for index in range(len(fitted_parts) - 1, 0, -1):
        if current_length() <= max_length:
            break
        component = fitted_parts[index]
        reduction = min(max(0, len(component) - 28), max(0, current_length() - max_length))
        if reduction:
            fitted_parts[index] = component[: len(component) - reduction].rstrip(" ._")
    return fitted_parts, fitted_file


def _windows_path_key(value: str) -> str:
    """Clave equivalente a la forma en que Windows compara nombres."""
    parts = []
    for part in PurePosixPath(value.replace("\\", "/").strip("/")).parts:
        normalized = unicodedata.normalize("NFC", part).rstrip(" .").casefold()
        parts.append(normalized)
    return "/".join(parts)


def _write_directory_chain(
    archive: zipfile.ZipFile,
    parts: list[str],
    written_directories: set[str],
) -> None:
    """Escribe explícitamente cada carpeta de la ruta sin duplicados de Windows."""

    for depth in range(1, len(parts) + 1):
        directory = "/".join(parts[:depth]).rstrip("/") + "/"
        directory_key = _windows_path_key(directory)
        if directory_key in written_directories:
            continue
        written_directories.add(directory_key)
        archive.writestr(directory, b"")


def _entry_folder_parts(entry: dict[str, Any], year: str) -> list[str]:
    dependency_code = str(entry.get("dependency_code", ""))
    series_code = str(entry.get("series_code", ""))
    code = str(entry.get("code", ""))
    parts = [
        _folder_component(year, "SIN_ANIO", max_length=12),
        _folder_component(
            f"{dependency_code} - {entry.get('dependency_name', '')}",
            "DEPENDENCIA",
            max_length=75,
        ),
        _folder_component(
            f"{dependency_code}.{series_code} - {entry.get('series_name', '')}",
            "SERIE",
            max_length=85,
        ),
    ]
    if entry.get("subseries_code"):
        parts.append(
            _folder_component(
                f"{code} - {entry.get('subseries_name', '')}",
                code or "SUBSERIE",
                max_length=95,
            )
        )
    return parts


def to_pdf_bytes(rows: list[dict[str, Any]], title: str = "Tabla de clasificación documental") -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = io.BytesIO()
    page_size = landscape(A4)
    document = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        rightMargin=8 * mm,
        leftMargin=8 * mm,
        topMargin=17 * mm,
        bottomMargin=14 * mm,
        title=title,
        author="Clasificador documental TRD/CCD PROINSALUD",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=15,
        leading=18,
        textColor=colors.HexColor("#123F73"),
        alignment=TA_CENTER,
        spaceAfter=4 * mm,
    )
    body_style = ParagraphStyle(
        "Cell",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=6.7,
        leading=8.1,
        textColor=colors.HexColor("#1F2933"),
    )
    header_style = ParagraphStyle(
        "HeaderCell",
        parent=body_style,
        fontName="Helvetica-Bold",
        textColor=colors.white,
        alignment=TA_CENTER,
    )

    def paragraph(value: Any, style: ParagraphStyle = body_style) -> Paragraph:
        text = str(value or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        return Paragraph(text, style)

    headers = ["Año", "Archivo", "Código final", "Estado", "Dependencia", "Serie / Subserie", "Afinidad"]
    table_data = [[paragraph(value, header_style) for value in headers]]
    for row in rows:
        series = str(row.get("serie", ""))
        if row.get("subserie"):
            series = f"{series} / {row.get('subserie')}"
        table_data.append(
            [
                paragraph(row.get("anio")),
                paragraph(row.get("archivo")),
                paragraph(row.get("codigo_final") or "Sin código"),
                paragraph(row.get("estado")),
                paragraph(row.get("dependencia")),
                paragraph(series),
                paragraph(f"{row.get('puntaje_similitud', 0)}%"),
            ]
        )

    table = Table(
        table_data,
        colWidths=[13 * mm, 52 * mm, 25 * mm, 23 * mm, 39 * mm, 94 * mm, 20 * mm],
        repeatRows=1,
        hAlign="CENTER",
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#123F73")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (2, 1), (3, -1), "CENTER"),
                ("ALIGN", (6, 1), (6, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#B8C7D6")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF4FA")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )

    generated = datetime.now().strftime("%Y-%m-%d %H:%M")

    def draw_page(canvas: Any, doc: Any) -> None:
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#E57900"))
        canvas.setLineWidth(0.9)
        canvas.line(8 * mm, page_size[1] - 10 * mm, page_size[0] - 8 * mm, page_size[1] - 10 * mm)
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#52606D"))
        canvas.drawString(8 * mm, 7 * mm, f"Generado: {generated}")
        canvas.drawRightString(page_size[0] - 8 * mm, 7 * mm, f"Página {doc.page}")
        canvas.restoreState()

    story = [Paragraph(title, title_style), Paragraph(f"Registros incluidos: {len(rows)}", body_style), Spacer(1, 3 * mm), table]
    document.build(story, onFirstPage=draw_page, onLaterPages=draw_page)
    return buffer.getvalue()


def _normalize(value: Any) -> str:
    text = str(value or "").lower()
    text = "".join(c for c in unicodedata.normalize("NFKD", text) if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _safe_template_directories(template_zip_path: str | Path | None) -> tuple[list[str], str]:
    if not template_zip_path or not Path(template_zip_path).is_file():
        return [], "CARPETAS_MASIVAS_ACTUALIZADAS"
    directories: set[str] = set()
    with zipfile.ZipFile(template_zip_path) as source:
        for info in source.infolist():
            raw = info.filename.replace("\\", "/").strip("/")
            if not raw or raw.startswith("/") or ".." in PurePosixPath(raw).parts:
                continue
            if info.is_dir() or info.filename.endswith("/"):
                directories.add(raw + "/")
    roots = [PurePosixPath(path.rstrip("/")).parts[0] for path in directories if PurePosixPath(path.rstrip("/")).parts]
    root = max(set(roots), key=roots.count) if roots else "CARPETAS_MASIVAS_ACTUALIZADAS"
    return sorted(directories, key=lambda value: (len(PurePosixPath(value.rstrip("/")).parts), value)), root


def _template_indexes(directories: list[str]) -> tuple[dict[str, list[str]], dict[str, list[str]]]:
    by_code: dict[str, list[str]] = {}
    children: dict[str, list[str]] = {}
    for directory in directories:
        parts = list(PurePosixPath(directory.rstrip("/")).parts)
        if not parts:
            continue
        component = parts[-1]
        match = re.match(r"^(\d+(?:\.\d+)*)\s*-\s*", component)
        if match:
            code = match.group(1)
            previous = by_code.get(code)
            if previous is None or len(parts) > len(previous):
                by_code[code] = parts
        if len(parts) > 1:
            parent = "/".join(parts[:-1])
            children.setdefault(parent, []).append(component)
    return by_code, children


def _similarity(query: str, candidate: str) -> float:
    q = _normalize(query)
    c = _normalize(candidate)
    if not q or not c:
        return 0.0
    if c in q:
        return 1.0
    q_tokens = set(q.split())
    c_tokens = set(c.split())
    overlap = len(q_tokens & c_tokens) / max(1, len(c_tokens))
    sequence = SequenceMatcher(None, q[:600], c).ratio()
    return max(overlap, sequence * 0.75)


def _choose_document_type(record: dict[str, Any], entry: dict[str, Any], options: list[str]) -> str:
    clean_options = [option for option in options if not re.match(r"^\d+(?:\.\d+)*\s*-", option)]
    if not clean_options:
        return ""
    if len(clean_options) == 1:
        return clean_options[0]
    text = f"{record.get('filename', '')} {str(record.get('extracted_text', ''))[:8_000]}"
    ranked = sorted(((option, _similarity(text, option)) for option in clean_options), key=lambda item: item[1], reverse=True)
    if ranked and ranked[0][1] >= 0.42:
        return ranked[0][0]

    # Como respaldo, cruza las denominaciones de la TRD con las carpetas disponibles,
    # pero no mezcla todos los tipos en la consulta porque generaría empates artificiales.
    documented = entry.get("document_types") or []
    if isinstance(documented, str):
        documented = [documented]
    fallback = []
    for option in clean_options:
        score = max((_similarity(str(document_type), option) for document_type in documented), default=0.0)
        fallback.append((option, score))
    fallback.sort(key=lambda item: item[1], reverse=True)
    return fallback[0][0] if fallback and fallback[0][1] >= 0.72 else ""


def _derived_massive_parts(entry: dict[str, Any], root: str) -> list[str]:
    dep_code = str(entry.get("dependency_code", ""))
    dep_name = str(entry.get("dependency_name", ""))
    if entry.get("is_other"):
        dependency = _folder_component(
            f"{dep_code} - {dep_name}" if dep_code else dep_name,
            OTHER_DEPENDENCY_FOLDER,
            80,
        )
        series_name = str(entry.get("series_name") or "").strip()
        subseries_name = str(entry.get("subseries_name") or "").strip()
        generic_other = not series_name or series_name.startswith("OTROS / OTRAS SERIES")
        if generic_other:
            return [root, dependency, OTHER_SERIES_FOLDER]
        parts = [
            root,
            dependency,
            _folder_component(f"SERIE SIN CÓDIGO - {series_name}", "SERIE SIN CÓDIGO", 110),
        ]
        if subseries_name:
            parts.append(
                _folder_component(
                    f"SUBSERIE SIN CÓDIGO - {subseries_name}",
                    "SUBSERIE SIN CÓDIGO",
                    120,
                )
            )
        return parts
    if not dep_code:
        return [root, OTHER_DEPENDENCY_FOLDER, OTHER_SERIES_FOLDER]
    series_code = str(entry.get("series_code", ""))
    code = str(entry.get("code", ""))
    parts = [
        root,
        _folder_component(f"{dep_code} - {dep_name}", dep_code, 80),
        _folder_component(f"{dep_code}.{series_code} - {entry.get('series_name', '')}", "SERIE", 90),
    ]
    if entry.get("subseries_code"):
        parts.append(_folder_component(f"{code} - {entry.get('subseries_name', '')}", code or "SUBSERIE", 100))
    return parts


def _massive_target_parts(
    record: dict[str, Any],
    entry: dict[str, Any],
    by_code: dict[str, list[str]],
    children: dict[str, list[str]],
    root: str,
    organize_by_document_type: bool,
) -> tuple[list[str], bool]:
    if record.get("status") == "Descartada":
        return [root, "DESCARTADOS"], False
    if entry.get("is_other"):
        return _derived_massive_parts(entry, root), False
    code = str(entry.get("code") or record.get("final_code") or record.get("suggested_code") or "")
    template_parts = by_code.get(code)
    parts = list(template_parts or _derived_massive_parts(entry, root))
    preserve_template_path = bool(template_parts)
    if organize_by_document_type and not entry.get("is_other"):
        parent = "/".join(parts)
        existing_children = children.get(parent, [])
        type_name = _choose_document_type(record, entry, existing_children)
        if not type_name and not existing_children:
            documented = entry.get("document_types") or []
            if isinstance(documented, str):
                documented = [documented]
            type_name = _choose_document_type(record, entry, [str(value) for value in documented])
        if type_name:
            # Si el tipo ya existe en la plantilla, se reutiliza exactamente el
            # nombre original; nunca se crea una variante recortada.
            exact_type = next(
                (candidate for candidate in existing_children if _windows_path_key(candidate) == _windows_path_key(type_name)),
                type_name,
            )
            parts.append(_folder_component(exact_type, "TIPO_DOCUMENTAL", 130))
            preserve_template_path = preserve_template_path and bool(existing_children)
    return parts, preserve_template_path


def build_classified_zip(
    records: list[dict[str, Any]],
    entries: list[dict[str, Any]],
    include_full_structure: bool = False,
    years: list[str] | None = None,
    dependency_codes: list[str] | None = None,
    structure_mode: str = "year",
    template_zip_path: str | Path | None = None,
    organize_by_document_type: bool = True,
) -> bytes:
    """Crea un ZIP por jerarquía anual o sobre la plantilla codificada de carpetas masivas."""
    if structure_mode not in {"year", "massive"}:
        raise ValueError("Modo de estructura no válido.")

    catalog = _catalog_map(entries)
    rows = manifest_rows(records, entries)
    buffer = io.BytesIO()
    used_paths: set[str] = set()
    written_directories: set[str] = set()
    missing_files: list[str] = []

    template_directories: list[str] = []
    template_root = "CARPETAS_MASIVAS_ACTUALIZADAS"
    by_code: dict[str, list[str]] = {}
    children: dict[str, list[str]] = {}
    if structure_mode == "massive":
        template_directories, template_root = _safe_template_directories(template_zip_path)
        by_code, children = _template_indexes(template_directories)

    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as archive:
        archive.writestr("MANIFIESTO_CLASIFICACION.csv", to_csv_bytes(rows))
        archive.writestr("MANIFIESTO_CLASIFICACION.json", to_json_bytes(rows))
        archive.writestr("TABLA_CLASIFICACION.pdf", to_pdf_bytes(rows))

        selected_dependencies = set(dependency_codes or [])
        if include_full_structure and structure_mode == "massive" and template_directories:
            for directory in template_directories:
                parts = PurePosixPath(directory.rstrip("/")).parts
                if selected_dependencies and len(parts) > 1:
                    dep_match = re.match(r"^(\d+(?:\.\d+)*)\s*-", parts[1])
                    if dep_match and dep_match.group(1) not in selected_dependencies:
                        continue
                directory_key = _windows_path_key(directory)
                if directory_key in written_directories:
                    continue
                written_directories.add(directory_key)
                archive.writestr(directory, b"")
        elif include_full_structure:
            selected_years = years or sorted({str(r.get("document_year") or "SIN_ANIO") for r in records})
            directory_paths: set[str] = set()
            for year in selected_years:
                for entry in entries:
                    if selected_dependencies and str(entry.get("dependency_code")) not in selected_dependencies:
                        continue
                    directory_parts, _ = _fit_zip_target(_entry_folder_parts(entry, year), "", max_length=220)
                    directory_paths.add("/".join(directory_parts) + "/")
            for directory in sorted(directory_paths):
                directory_key = _windows_path_key(directory)
                if directory_key in written_directories:
                    continue
                written_directories.add(directory_key)
                archive.writestr(directory, b"")

        for record in records:
            source = Path(str(record.get("stored_path", "")))
            if not source.is_file():
                missing_files.append(str(record.get("filename", source.name)))
                continue
            entry = canonicalize_custom_entry(
                _effective_entry(record, catalog),
                entries,
                filename=str(record.get("filename") or ""),
                extracted_text=str(record.get("extracted_text") or ""),
            )
            preserve_template_path = False
            if structure_mode == "massive":
                parts, preserve_template_path = _massive_target_parts(
                    record,
                    entry,
                    by_code,
                    children,
                    template_root,
                    organize_by_document_type,
                )
            else:
                year = str(record.get("document_year") or "SIN_ANIO")
                if record.get("status") == "Descartada":
                    parts = [_folder_component(year, "SIN_ANIO"), "DESCARTADOS"]
                elif entry and entry.get("code"):
                    parts = _entry_folder_parts(entry, year)
                elif entry.get("is_other"):
                    dependency_code = str(entry.get("dependency_code") or "").strip()
                    dependency_name = str(entry.get("dependency_name") or "").strip()
                    dependency_value = (
                        f"{dependency_code} - {dependency_name}"
                        if dependency_code or dependency_name
                        else ""
                    )
                    dep = _folder_component(
                        dependency_value,
                        OTHER_DEPENDENCY_FOLDER,
                    )
                    parts = [_folder_component(year, "SIN_ANIO"), dep, OTHER_SERIES_FOLDER]
                else:
                    parts = [_folder_component(year, "SIN_ANIO"), OTHER_DEPENDENCY_FOLDER, OTHER_SERIES_FOLDER]

            filename = _file_component(Path(str(record.get("filename", source.name))).name, source.name)
            parts, filename = _fit_zip_target(
                parts,
                filename,
                max_length=300 if preserve_template_path else 240,
                preserve_parts=preserve_template_path,
            )
            target = "/".join(parts + [filename])
            stem = Path(filename).stem
            suffix = Path(filename).suffix
            counter = 2
            target_key = _windows_path_key(target)
            while target_key in used_paths:
                target = "/".join(parts + [f"{stem}_{counter}{suffix}"])
                target_key = _windows_path_key(target)
                counter += 1
            used_paths.add(target_key)
            _write_directory_chain(archive, parts, written_directories)
            archive.write(source, target)

        if missing_files:
            archive.writestr(
                "ERRORES_EXPORTACION.txt",
                "No se encontraron los archivos originales de los siguientes registros:\n"
                + "\n".join(f"- {name}" for name in missing_files),
            )
    return buffer.getvalue()
