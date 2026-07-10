from __future__ import annotations

import csv
import io
import json
import re
import unicodedata
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


TABLE_EXTENSIONS = {".xlsx", ".xlsm", ".xls", ".csv", ".tsv"}
ARCHIVE_EXTENSIONS = {".zip", ".rar"}
MAX_ARCHIVE_MEMBERS = 1_000
MAX_MEMBER_BYTES = 50 * 1024 * 1024
MAX_ARCHIVE_BYTES = 250 * 1024 * 1024


@dataclass
class CatalogBuildResult:
    entries: list[dict[str, Any]] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    sources: list[str] = field(default_factory=list)
    trd_rows: int = 0
    ccd_rows: int = 0

    @property
    def stats(self) -> dict[str, int]:
        return {
            "registros": len(self.entries),
            "dependencias": len({e.get("dependency_code") for e in self.entries if e.get("dependency_code")}),
            "series": len({(e.get("dependency_code"), e.get("series_code")) for e in self.entries}),
            "subseries": sum(bool(e.get("subseries_code")) for e in self.entries),
            "fuentes": len(set(self.sources)),
        }


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").replace("\r", " ").replace("\n", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize_for_search(value: Any) -> str:
    text = clean_text(value).lower()
    text = "".join(c for c in unicodedata.normalize("NFKD", text) if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9.]+", " ", text).strip()


def normalize_code(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = clean_text(value).replace(",", ".")
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"\.0$", "", text)
    return text if re.fullmatch(r"\d+(?:\.\d+)*", text) else ""


def _name_from_upload(item: Any) -> tuple[str, bytes]:
    if isinstance(item, tuple) and len(item) == 2:
        return str(item[0]), bytes(item[1])
    if isinstance(item, dict):
        return str(item["name"]), bytes(item["data"])
    name = getattr(item, "name", "archivo")
    if hasattr(item, "getvalue"):
        return str(name), bytes(item.getvalue())
    if hasattr(item, "read"):
        return str(name), bytes(item.read())
    raise TypeError("Cada archivo debe tener nombre y contenido binario.")


def _safe_archive_name(name: str) -> str:
    normalized = name.replace("\\", "/").lstrip("/")
    parts = [p for p in normalized.split("/") if p not in ("", ".", "..")]
    return "/".join(parts)


def _expand_archive(name: str, data: bytes) -> tuple[list[tuple[str, bytes]], list[str]]:
    files: list[tuple[str, bytes]] = []
    errors: list[str] = []
    ext = Path(name).suffix.lower()
    total = 0

    if ext == ".zip":
        try:
            with zipfile.ZipFile(io.BytesIO(data)) as archive:
                members = [m for m in archive.infolist() if not m.is_dir()]
                if len(members) > MAX_ARCHIVE_MEMBERS:
                    raise ValueError(f"El ZIP contiene más de {MAX_ARCHIVE_MEMBERS} archivos.")
                for member in members:
                    member_name = _safe_archive_name(member.filename)
                    member_ext = Path(member_name).suffix.lower()
                    if member_ext not in TABLE_EXTENSIONS:
                        continue
                    if member.file_size > MAX_MEMBER_BYTES:
                        errors.append(f"{member_name}: supera 50 MB y fue omitido.")
                        continue
                    total += member.file_size
                    if total > MAX_ARCHIVE_BYTES:
                        raise ValueError("El contenido descomprimido supera 250 MB.")
                    files.append((f"{name}/{member_name}", archive.read(member)))
        except Exception as exc:
            errors.append(f"{name}: no se pudo leer el ZIP ({exc}).")
        return files, errors

    if ext == ".rar":
        try:
            import rarfile

            with rarfile.RarFile(io.BytesIO(data)) as archive:
                members = [m for m in archive.infolist() if not m.isdir()]
                if len(members) > MAX_ARCHIVE_MEMBERS:
                    raise ValueError(f"El RAR contiene más de {MAX_ARCHIVE_MEMBERS} archivos.")
                for member in members:
                    member_name = _safe_archive_name(member.filename)
                    member_ext = Path(member_name).suffix.lower()
                    if member_ext not in TABLE_EXTENSIONS:
                        continue
                    if member.file_size > MAX_MEMBER_BYTES:
                        errors.append(f"{member_name}: supera 50 MB y fue omitido.")
                        continue
                    total += member.file_size
                    if total > MAX_ARCHIVE_BYTES:
                        raise ValueError("El contenido descomprimido supera 250 MB.")
                    files.append((f"{name}/{member_name}", archive.read(member)))
        except Exception as exc:
            errors.append(
                f"{name}: no se pudo abrir el RAR ({exc}). Instale 7-Zip/UnRAR o conviértalo a ZIP."
            )
        return files, errors

    return files, [f"{name}: formato de archivo comprimido no compatible."]


def expand_table_uploads(items: Iterable[Any]) -> tuple[list[tuple[str, bytes]], list[str]]:
    expanded: list[tuple[str, bytes]] = []
    errors: list[str] = []
    for item in items:
        name, data = _name_from_upload(item)
        ext = Path(name).suffix.lower()
        if ext in TABLE_EXTENSIONS:
            expanded.append((name, data))
        elif ext in ARCHIVE_EXTENSIONS:
            nested, nested_errors = _expand_archive(name, data)
            expanded.extend(nested)
            errors.extend(nested_errors)
        else:
            errors.append(f"{name}: tipo de tabla no compatible.")
    return expanded, errors


def _workbook_kind(ws: Any) -> str:
    values = " ".join(clean_text(ws.cell(r, c).value).upper() for r in range(1, min(ws.max_row, 12) + 1) for c in range(1, min(ws.max_column, 12) + 1))
    if "TABLA DE RETENCIÓN DOCUMENTAL" in values or "TABLA DE RETENCION DOCUMENTAL" in values:
        return "TRD"
    if "CUADRO DE CLASIFICACIÓN DOCUMENTAL" in values or "CUADRO DE CLASIFICACION DOCUMENTAL" in values:
        return "CCD"
    headers = {normalize_for_search(ws.cell(r, c).value) for r in range(1, min(ws.max_row, 20) + 1) for c in range(1, min(ws.max_column, 15) + 1)}
    if "series" in headers and "subseries" in headers and "seccion" in headers:
        return "CCD"
    if "retencion" in headers and any("serie" in h for h in headers):
        return "TRD"
    return "UNKNOWN"


def _find_header_row(ws: Any, expected: str = "codigo") -> int:
    for row in range(1, min(ws.max_row, 40) + 1):
        first = normalize_for_search(ws.cell(row, 1).value)
        row_text = " ".join(normalize_for_search(ws.cell(row, c).value) for c in range(1, min(ws.max_column, 15) + 1))
        if first == expected and ("serie" in row_text or "seccion" in row_text):
            return row
    return 5


def _dependency_from_trd(ws: Any, source_name: str) -> tuple[str, str]:
    dep_name = ""
    dep_code = ""
    for row in range(1, min(ws.max_row, 20) + 1):
        for col in range(1, min(ws.max_column, 15) + 1):
            label = normalize_for_search(ws.cell(row, col).value)
            if "dependencia productora" in label:
                for next_col in range(col + 1, min(ws.max_column, col + 4) + 1):
                    candidate = clean_text(ws.cell(row, next_col).value)
                    if candidate:
                        dep_name = candidate
                        break
            if label in {"codigo", "codigo dependencia"}:
                for next_col in range(col + 1, min(ws.max_column, col + 4) + 1):
                    candidate = normalize_code(ws.cell(row, next_col).value)
                    if candidate:
                        dep_code = candidate
    file_match = re.search(r"(?:^|/)(\d+(?:\.\d+)*)\s+TRD\s+(.+?)\.(?:xlsx|xlsm|xls)$", source_name, re.I)
    if file_match:
        dep_code = dep_code or file_match.group(1)
        dep_name = dep_name or file_match.group(2)
    return dep_code, dep_name


def _split_code(full_code: str, dependency_code: str) -> tuple[str, str]:
    full_parts = full_code.split(".")
    dep_parts = dependency_code.split(".") if dependency_code else []
    if dep_parts and full_parts[: len(dep_parts)] == dep_parts:
        remainder = full_parts[len(dep_parts) :]
    else:
        remainder = full_parts[1:]
    series_code = remainder[0] if remainder else ""
    subseries_code = ".".join(remainder[1:]) if len(remainder) > 1 else ""
    return series_code, subseries_code


def _trd_columns(ws: Any, header_row: int) -> dict[str, int]:
    columns = {
        "paper": 3,
        "electronic": 4,
        "management": 5,
        "central": 6,
        "conservation": 7,
        "selection": 8,
        "elimination": 9,
        "digitalization": 10,
        "procedure": 11,
    }
    subheader_row = header_row + 1
    for col in range(1, ws.max_column + 1):
        header = normalize_for_search(ws.cell(header_row, col).value)
        subheader = normalize_for_search(ws.cell(subheader_row, col).value)
        if "procedimiento" in header:
            columns["procedure"] = col
        if subheader == "papel":
            columns["paper"] = col
        elif "electronico" in subheader:
            columns["electronic"] = col
        elif "archivo de gestion" in subheader:
            columns["management"] = col
        elif "archivo central" in subheader:
            columns["central"] = col
        elif subheader == "c":
            columns["conservation"] = col
        elif subheader == "s":
            columns["selection"] = col
        elif subheader == "e":
            columns["elimination"] = col
        elif subheader == "d":
            columns["digitalization"] = col
    return columns


def _disposition_from_row(ws: Any, row: int, columns: dict[str, int]) -> list[str]:
    labels = {
        "conservation": "Conservación total",
        "selection": "Selección",
        "elimination": "Eliminación",
        "digitalization": "Digitalización/Reproducción",
    }
    return [
        label
        for key, label in labels.items()
        if clean_text(ws.cell(row, columns[key]).value).upper() == "X"
    ]


def parse_trd_sheet(ws: Any, source_name: str) -> list[dict[str, Any]]:
    header_row = _find_header_row(ws)
    columns = _trd_columns(ws, header_row)
    dep_code, dep_name = _dependency_from_trd(ws, source_name)
    if not dep_code:
        return []

    anchors: list[int] = []
    for row in range(header_row + 1, ws.max_row + 1):
        code = normalize_code(ws.cell(row, 1).value)
        if code and (code == dep_code or code.startswith(dep_code + ".")):
            anchors.append(row)

    entries: list[dict[str, Any]] = []
    for index, row in enumerate(anchors):
        next_row = anchors[index + 1] if index + 1 < len(anchors) else ws.max_row + 1
        code = normalize_code(ws.cell(row, 1).value)
        series_code, subseries_code = _split_code(code, dep_code)
        series_name = clean_text(ws.cell(row, 2).value)
        if not series_code or not series_name:
            continue

        children: list[tuple[int, str, str, str]] = []
        for child_row in range(row + 1, next_row):
            label = clean_text(ws.cell(child_row, 2).value)
            if not label:
                continue
            support_paper = clean_text(ws.cell(child_row, columns["paper"]).value)
            support_electronic = clean_text(ws.cell(child_row, columns["electronic"]).value)
            children.append((child_row, label, support_paper, support_electronic))

        subseries_name = ""
        if subseries_code and children:
            subseries_name = children[0][1]
            children = children[1:]

        document_types = []
        supports = []
        for _, label, paper, electronic in children:
            normalized_label = normalize_for_search(label)
            if normalized_label.startswith(("convenciones", "firmas responsables", "elaborado por", "aprobado por")):
                break
            document_types.append(label)
            support = ", ".join(x for x in ["Papel" if paper.upper() == "X" else paper, electronic] if x)
            supports.append(support)

        entries.append(
            {
                "code": code,
                "dependency_code": dep_code,
                "dependency_name": dep_name,
                "section_code": "",
                "section_name": "",
                "subsection_code": dep_code,
                "subsection_name": dep_name,
                "series_code": series_code,
                "series_name": series_name,
                "subseries_code": subseries_code,
                "subseries_name": subseries_name,
                "document_types": document_types,
                "supports": supports,
                "retention_management": clean_text(ws.cell(row, columns["management"]).value),
                "retention_central": clean_text(ws.cell(row, columns["central"]).value),
                "final_disposition": _disposition_from_row(ws, row, columns),
                "procedure": clean_text(ws.cell(row, columns["procedure"]).value),
                "source_type": "TRD",
                "source_files": [source_name],
                "source_sheets": [ws.title],
            }
        )
    return entries


def parse_ccd_sheet(ws: Any, source_name: str) -> list[dict[str, Any]]:
    header_row = _find_header_row(ws)
    section_code = section_name = subsection_code = subsection_name = ""
    series_code = series_name = ""
    entries: list[dict[str, Any]] = []

    for row in range(header_row + 1, ws.max_row + 1):
        a = normalize_code(ws.cell(row, 1).value)
        b = clean_text(ws.cell(row, 2).value)
        c = normalize_code(ws.cell(row, 3).value)
        d = clean_text(ws.cell(row, 4).value)
        e = normalize_code(ws.cell(row, 5).value)
        f = clean_text(ws.cell(row, 6).value)
        g = normalize_code(ws.cell(row, 7).value)
        h = clean_text(ws.cell(row, 8).value)

        if a or b:
            section_code = a or section_code
            section_name = b or section_name
            if not c and not d:
                subsection_code = subsection_name = ""
        if c or d:
            subsection_code = c or subsection_code
            subsection_name = d or subsection_name
        if e or f:
            series_code = e or series_code
            series_name = f or series_name

        dependency_code = subsection_code or section_code
        dependency_name = subsection_name or section_name
        if not dependency_code or not series_code or not series_name:
            continue
        if not (e or f or g or h):
            continue

        code = f"{dependency_code}.{series_code}"
        if g:
            code = f"{code}.{g}"
        entries.append(
            {
                "code": code,
                "dependency_code": dependency_code,
                "dependency_name": dependency_name,
                "section_code": section_code,
                "section_name": section_name,
                "subsection_code": subsection_code,
                "subsection_name": subsection_name,
                "series_code": series_code,
                "series_name": series_name,
                "subseries_code": g,
                "subseries_name": h,
                "document_types": [],
                "supports": [],
                "retention_management": "",
                "retention_central": "",
                "final_disposition": [],
                "procedure": "",
                "source_type": "CCD",
                "source_files": [source_name],
                "source_sheets": [ws.title],
            }
        )
    return entries


def _read_delimited(name: str, data: bytes) -> list[dict[str, Any]]:
    try:
        from charset_normalizer import from_bytes

        match = from_bytes(data).best()
        text = str(match) if match else data.decode("utf-8", errors="replace")
    except Exception:
        text = data.decode("utf-8", errors="replace")
    delimiter = "\t" if Path(name).suffix.lower() == ".tsv" else ","
    sample = text[:4096]
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        pass
    rows = list(csv.DictReader(io.StringIO(text), delimiter=delimiter))
    normalized: list[dict[str, Any]] = []
    for raw in rows:
        mapping = {normalize_for_search(key): clean_text(value) for key, value in raw.items() if key}
        dep_code = mapping.get("codigo dependencia") or mapping.get("dependencia codigo") or mapping.get("codigo seccion") or ""
        dep_name = mapping.get("dependencia") or mapping.get("subseccion") or mapping.get("seccion") or ""
        series_code = mapping.get("codigo serie") or ""
        series_name = mapping.get("serie") or mapping.get("series") or ""
        subseries_code = mapping.get("codigo subserie") or ""
        subseries_name = mapping.get("subserie") or mapping.get("subseries") or ""
        code = mapping.get("codigo") or ".".join(x for x in [dep_code, series_code, subseries_code] if x)
        if code and series_name:
            normalized.append(
                {
                    "code": code,
                    "dependency_code": dep_code,
                    "dependency_name": dep_name,
                    "section_code": "",
                    "section_name": "",
                    "subsection_code": dep_code,
                    "subsection_name": dep_name,
                    "series_code": series_code,
                    "series_name": series_name,
                    "subseries_code": subseries_code,
                    "subseries_name": subseries_name,
                    "document_types": [x.strip() for x in mapping.get("tipos documentales", "").split("|") if x.strip()],
                    "supports": [],
                    "retention_management": mapping.get("retencion gestion", ""),
                    "retention_central": mapping.get("retencion central", ""),
                    "final_disposition": [x.strip() for x in mapping.get("disposicion final", "").split("|") if x.strip()],
                    "procedure": mapping.get("procedimiento", ""),
                    "source_type": "TABLA",
                    "source_files": [name],
                    "source_sheets": ["datos"],
                }
            )
    return normalized


def _read_legacy_excel(name: str, data: bytes) -> list[dict[str, Any]]:
    import pandas as pd

    class CellValue:
        def __init__(self, value: Any) -> None:
            self.value = None if pd.isna(value) else value

    class DataFrameSheet:
        def __init__(self, title: str, frame: Any) -> None:
            self.title = title
            self.frame = frame
            self.max_row = int(frame.shape[0])
            self.max_column = int(frame.shape[1])

        def cell(self, row: int, column: int) -> CellValue:
            if row < 1 or column < 1 or row > self.max_row or column > self.max_column:
                return CellValue(None)
            return CellValue(self.frame.iat[row - 1, column - 1])

    entries: list[dict[str, Any]] = []
    workbook = pd.ExcelFile(io.BytesIO(data))
    for sheet_name in workbook.sheet_names:
        frame = workbook.parse(sheet_name=sheet_name, header=None)
        sheet = DataFrameSheet(sheet_name, frame)
        kind = _workbook_kind(sheet)
        if kind == "TRD":
            entries.extend(parse_trd_sheet(sheet, name))
        elif kind == "CCD":
            entries.extend(parse_ccd_sheet(sheet, name))
    return entries


def merge_catalog_entries(entries: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for incoming in entries:
        code = normalize_code(incoming.get("code"))
        if not code:
            continue
        incoming = dict(incoming)
        incoming["code"] = code
        current = grouped.get(code)
        if current is None:
            grouped[code] = incoming
            continue

        current_is_trd = current.get("source_type") == "TRD"
        incoming_is_trd = incoming.get("source_type") == "TRD"
        for field in [
            "dependency_code",
            "dependency_name",
            "section_code",
            "section_name",
            "subsection_code",
            "subsection_name",
            "series_code",
            "series_name",
            "subseries_code",
            "subseries_name",
        ]:
            if incoming.get(field) and (not current.get(field) or (not incoming_is_trd and field.endswith("name"))):
                current[field] = incoming[field]

        for field in ["retention_management", "retention_central", "procedure"]:
            if incoming.get(field) and (not current.get(field) or incoming_is_trd):
                current[field] = incoming[field]
        for field in ["document_types", "supports", "final_disposition", "source_files", "source_sheets"]:
            values = [clean_text(x) for x in (current.get(field) or []) + (incoming.get(field) or [])]
            current[field] = list(dict.fromkeys(x for x in values if x))
        if incoming_is_trd or not current_is_trd:
            current["source_type"] = "TRD" if incoming_is_trd else current.get("source_type", "CCD")

    result = list(grouped.values())
    for entry in result:
        entry["search_text"] = " ".join(
            x
            for x in [
                entry.get("code", ""),
                entry.get("dependency_code", ""),
                entry.get("dependency_name", ""),
                entry.get("series_name", ""),
                entry.get("subseries_name", ""),
                " ".join(entry.get("document_types") or []),
                entry.get("procedure", ""),
            ]
            if x
        )
    return sorted(result, key=lambda e: tuple(int(p) for p in e["code"].split(".")))


def build_catalog_from_files(items: Iterable[Any]) -> CatalogBuildResult:
    result = CatalogBuildResult()
    files, errors = expand_table_uploads(items)
    result.errors.extend(errors)
    all_entries: list[dict[str, Any]] = []

    for name, data in files:
        result.sources.append(name)
        ext = Path(name).suffix.lower()
        try:
            if ext in {".csv", ".tsv"}:
                rows = _read_delimited(name, data)
                all_entries.extend(rows)
                continue
            if ext == ".xls":
                rows = _read_legacy_excel(name, data)
                all_entries.extend(rows)
                continue

            workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=False, keep_vba=False)
            for ws in workbook.worksheets:
                kind = _workbook_kind(ws)
                if kind == "TRD":
                    rows = parse_trd_sheet(ws, name)
                    result.trd_rows += len(rows)
                    all_entries.extend(rows)
                elif kind == "CCD":
                    rows = parse_ccd_sheet(ws, name)
                    result.ccd_rows += len(rows)
                    all_entries.extend(rows)
        except Exception as exc:
            result.errors.append(f"{name}: error al interpretar la tabla ({exc}).")

    result.entries = merge_catalog_entries(all_entries)
    if not result.entries and not result.errors:
        result.errors.append("No se encontraron filas reconocibles de TRD o CCD.")
    return result


def save_catalog_json(entries: list[dict[str, Any]], path: str | Path, metadata: dict[str, Any] | None = None) -> None:
    payload = {"metadata": metadata or {}, "entries": entries}
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    Path(path).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def load_catalog_json(path: str | Path) -> list[dict[str, Any]]:
    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    return payload["entries"] if isinstance(payload, dict) else payload
